import os
import random
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


def _read_interviewee_sheet():
    """
        Check the bellow JS function to get the schedule from when2meet.
        (https://www.when2meet.com/?18687020-njdRB)
        function getCSV() {
          result = "Time," + PeopleNames.join(",")+"\n";
          for(let i = 0; i < AvailableAtSlot.length; i++) {
              let slot = $x(`string(//div[@id="GroupTime${TimeOfSlot[i]}"]/@onmouseover)`);
              slot = slot.match(/.*"(.*)".*/)[1];
              result += slot + ",";
              result += PeopleIDs.map(id => AvailableAtSlot[i].includes(id) ? 1 : 0).join(",");
              result+= "\n";
          }
          console.log(result);
        }
        getCSV();
        :return:
        """
    sheet = defaultdict(list)
    interviewee = "/".join([os.getcwd(), "input", "interviewee.txt"])
    with open(interviewee, "r") as f:
        date = 25
        for interviewee_row in f.readlines():
            if "#" in interviewee_row:
                date += 1
            else:
                # formatting
                avaliables = interviewee_row.split(' - ')
                avaliables[1] = avaliables[1].replace("\'", "")
                avaliables[1] = avaliables[1].replace(", ", ",")
                avaliables[1] = avaliables[1][1:-2]
                avaliables[1] = avaliables[1].split(",")
                avaliables[0] = ":".join([str(date), avaliables[0].split(" ~ ")[0]])
                sheet[avaliables[0]].extend(avaliables[1])
    return sheet


def _read_interviewer_sheet():
    sheet = defaultdict(list)
    interviewer = "/".join([os.getcwd(), "input", "interviewer.csv"])
    df = pd.read_csv(interviewer)
    for row_id, row in df.iterrows():
        time_slot = datetime.strptime(row[0], '%a %d %b %Y %I:%M:%S %p %Z')
        time_slot = time_slot.strftime('%d:%H:%M')
        for co_id, interviewer in enumerate(row):
            if interviewer == 1:
                sheet[time_slot].append(df.columns[co_id])
    return sheet


def generate_interview_schedule():
    interviewers = _read_interviewee_sheet()
    interviewees = _read_interviewer_sheet()

    candidates = dict()
    assignments = list()
    all_times = list(set(interviewers.keys()) & set(interviewees.keys()))

    for t in all_times:
        candidates[t] = {
            'interviewer': interviewers[t],
            'interviewee': interviewees[t]
        }
    for t in all_times:
        candidates_t = candidates[t]
        if len(candidates_t['interviewer']) > 0 and len(candidates_t['interviewee']) > 0:
            interviewee = random.choice(candidates_t['interviewee'])
            interviewers_t = candidates_t['interviewer']
            interviewer = random.choice(interviewers_t)
            interviewers_t.remove(interviewer)
            assignments.append({
                "date:time": t,
                "interviewer": interviewer,
                "interviewee": interviewee,
                "candidates": ", ".join(interviewers_t)
            })
            candidates_t['interviewee'].remove(interviewee)

    return assignments


def make_schedule_sheet(sheet):
    df = pd.DataFrame(sheet)
    df = df.sort_values(by=["date:time"], ascending=[True])
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    pink_fill = PatternFill(start_color='F781F3', end_color='F781F3', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_col=len(df.columns) + 1):
        for cell in row:
            # coloring test
            if '25' in str(cell.value):
                cell.fill = pink_fill
            elif '26' in str(cell.value):
                cell.fill = green_fill
    wb.save("output/schedule.xlsx")


if __name__ == "__main__":
    schedule_candidate = generate_interview_schedule()
    make_schedule_sheet(schedule_candidate)


